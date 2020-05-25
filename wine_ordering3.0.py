"""
Created on Mon May 11 17:21:19 2020

Purpose: To compile, generate purchase order and email orders from an Excel spreadsheet.

Author: Daniel Sharp
"""

import os
import openpyxl
import generate_po_ref as pr
import generate_date as gd
import generate_table as gt
import generate_pdf as gp
import is_empty as ie
import send_drafts as sd
import pandas as pd
import time

INCREMENT = 20 # Define order block length. Change to worksheet specifications. 

print('Hello!\n \nCompiling orders...\n')

wb = openpyxl.load_workbook('Order Sheet/order_sheet_2020.xlsx') # Open order workbook.
sheet = wb.active # Select active sheet
row_count = sheet.max_row # Get length of sheet

start = 3 # Start value of first order block.
end = 21 # End value of first order block.
cost_list = [] # Create empty cost list.
email_list = []
company_list = []
body_list = []
po_list = []
cc_list = []
bcc_list = []
drafts = {}

for order_blocks in range(0,row_count): # Select worksheet rows to iterate over.

    order = {} # Create empty order dictionary.
    units_list = [] # Create empty number of units list.
    product_list = [] # Create empty product list
    luc_list = [] # Create empty price list.
    
    for cell_value in range(start, end): # Iterate over cells in order block.
        date = gd.generate_date() # Call date function.
        po_ref = pr.generate_po_ref() # Call generate po refernce function.
        
        company = sheet.cell(start,1).value # Get company name from spreadsheet
        if ie.is_empty(company) == True: # If no company in cell, break out of loop.
            break
        
        full_name = sheet.cell(start + 1, 1).value # Get email recipient.
        mobile = sheet.cell(start + 2, 1).value # Get mobile number.
        email = sheet.cell(start + 3, 1).value # Get company email.
        cc_email = sheet.cell(start + 8, 1).value # Get cc email.
        bcc_email = sheet.cell(start + 10, 1).value # Get bcc email.
        filename = company[0:3].upper() + po_ref # Create unique PO reference.
        notes = sheet.cell(start + 16, 1).value # Get invoice notes.
        body = sheet.cell(start + 12, 1).value # Get email body.
        suffix = '.pdf'
        po = os.path.join('Draft Orders',date, company, filename + suffix) # Defines PO attachment name.
        
        quantity = sheet.cell(cell_value,3).value # Get quantity ordered.
        
        if ie.is_empty(quantity) == False: # Create order if a number of units is specified.
            
            product = sheet.cell(cell_value,2).value # Create list of products ordered.
            product_list.append(product)            
            
            units = sheet.cell(cell_value,3).value # Create list of units ordered.
            units_list.append(units)
            
            luc = sheet.cell(cell_value,4).value # Create list of prices.
            luc_list.append(luc)
            
        order['Units'] = units_list # Create order dictionary of lists.
        order['Product'] = product_list
        order['LUC'] = luc_list
    
    if ie.is_empty(units_list) == False: # If units list is not empty create order pdf and email.
        gt.generate_table(order) # Create order table and PO pdf.
        gp.generate_pdf(company, filename, full_name, mobile, email, date, notes)
        
        email_list.append(email)
        company_list.append(company)
        body_list.append(body)
        po_list.append(po)
        cc_list.append(cc_email)
        bcc_list.append(bcc_email)
        
        cost = gt.generate_table(order)
        cost_list.append(cost)
        
        sheet.cell(start + 4, 1).value = 'PO GENERATED: ' + date # Save date, PO ref and cost to spreadsheet.
        sheet.cell(start + 5, 1).value = 'PO REF: ' + filename
        sheet.cell(start + 6, 1).value = 'TOTAL: $' + cost
        
        cost = float(cost)
        
        drafts['email'] = email_list
        drafts['company'] = company_list
        drafts['body'] = body_list
        drafts['po'] = po_list
        drafts['cc_email'] = cc_list
        drafts['bcc_email'] = bcc_list
        drafts['cost'] = cost_list
        print('* * *\n')
        time.sleep(1)
        
    start += INCREMENT # Increment start and end values to next block.
    end += INCREMENT
    
wb.save('Order Sheet/order_sheet_2020.xlsx') # Save workbook

df = pd.DataFrame(drafts)
df['cost'] = df['cost'].astype(float)

num_orders = len(df.index)
total_cost = round(df['cost'].sum(),2)

if num_orders == 1:
    print(f'You have {num_orders} draft order with a total cost inc GST of: ${total_cost}.\n')
else:
    print(f'You have {num_orders} draft orders with a total cost inc GST of: ${total_cost}.\n')
    
for i,j in df.iterrows(): 
    print(i,j)
    print()

sd.send_drafts(df)


